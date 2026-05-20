import sqlite3
import openpyxl
import os
from pathlib import Path

# Caminhos locais
DB_PATH = Path("quadro_pessoal.db")
XLSX_PATH = Path("CARGOS EFETIVOS E COMISSIONADOS - PMM - Ativos e extintos - 2026.xlsx")

def migrar():
    print("Iniciando migração do banco de dados...")
    
    if not DB_PATH.exists():
        print(f"Erro: Banco de dados não encontrado em {DB_PATH.absolute()}")
        return

    # 1. Conectar e Adicionar Colunas
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Obter colunas existentes
    cursor.execute("PRAGMA table_info(Cargos)")
    colunas_existentes = [row[1] for row in cursor.fetchall()]
    
    novas_colunas = {
        "recrutamento": "TEXT",
        "restricao_exigencia": "TEXT",
        "fonte_carga_horaria": "TEXT",
        "fonte_atribuicoes": "TEXT"
    }
    
    for col_nome, col_tipo in novas_colunas.items():
        if col_nome not in colunas_existentes:
            print(f"Adicionando coluna '{col_nome}' ({col_tipo}) na tabela Cargos...")
            cursor.execute(f"ALTER TABLE Cargos ADD COLUMN {col_nome} {col_tipo}")
        else:
            print(f"Coluna '{col_nome}' já existe.")
            
    conn.commit()

    # 2. Copiar escolaridade para restricao_exigencia
    print("Copiando dados existentes de 'escolaridade' para 'restricao_exigencia'...")
    cursor.execute("UPDATE Cargos SET restricao_exigencia = escolaridade WHERE restricao_exigencia IS NULL")
    conn.commit()

    # 3. Recriar a View vw_SaldoVagas
    print("Recriando a view vw_SaldoVagas para incluir os novos campos...")
    cursor.execute("DROP VIEW IF EXISTS vw_SaldoVagas")
    cursor.execute("""
        CREATE VIEW vw_SaldoVagas AS
        SELECT
            id, nome, codigo_fopag, situacao, situacao_delib, tipo_provimento,
            escolaridade, carga_horaria, simbolo_vencimento,
            total_previstos, total_ocupados,
            (total_previstos - total_ocupados) AS saldo_vagas,
            CASE WHEN (total_previstos - total_ocupados) < 0 THEN 1 ELSE 0 END AS alerta_saldo_negativo,
            atribuicoes, criado_em, atualizado_em,
            recrutamento, restricao_exigencia, fonte_carga_horaria, fonte_atribuicoes
        FROM Cargos;
    """)
    conn.commit()

    # 4. Ler dados do Excel e preencher recrutamento/fonte_carga_horaria
    if XLSX_PATH.exists():
        print(f"Planilha encontrada: {XLSX_PATH.name}. Extraindo dados históricos...")
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        sheet = wb.active
        
        # Mapear cabeçalhos
        headers = [cell.value for cell in next(sheet.iter_rows())]
        try:
            idx_cargo = headers.index('Cargo')
            idx_fopag = headers.index('Código FOPAG')
            idx_recrut = headers.index('Recrutamento')
            idx_fonte = headers.index('Fonte Carga Horária')
        except ValueError as e:
            print(f"Erro: Colunas esperadas não encontradas no Excel. Detalhes: {e}")
            conn.close()
            return
            
        print("Mapeando linhas do Excel...")
        excel_rows = list(sheet.iter_rows())[1:]
        
        # Buscar todos os cargos do banco
        cursor.execute("SELECT id, nome, codigo_fopag FROM Cargos")
        db_cargos = cursor.fetchall()
        
        atualizados = 0
        for row in excel_rows:
            nome_val = row[idx_cargo].value
            fopag_val = row[idx_fopag].value
            recrut_val = row[idx_recrut].value
            fonte_val = row[idx_fonte].value
            
            if not nome_val:
                continue
                
            nome_clean = str(nome_val).strip().lower()
            fopag_clean = str(fopag_val).strip() if fopag_val is not None else None
            
            # Tentar combinar pelo FOPAG ou pelo nome
            cargo_id = None
            for db_id, db_nome, db_fopag in db_cargos:
                # Se código FOPAG coincidir
                if fopag_clean and db_fopag and str(db_fopag).strip() == fopag_clean:
                    cargo_id = db_id
                    break
                # Se nome coincidir
                if str(db_nome).strip().lower() == nome_clean:
                    cargo_id = db_id
                    break
                    
            if cargo_id:
                # Sanitizar valores do Excel
                recrut_clean = str(recrut_val).strip() if recrut_val is not None else None
                if recrut_clean == '-':
                    recrut_clean = None
                    
                fonte_clean = str(fonte_val).strip() if fonte_val is not None else None
                
                # Executar update
                cursor.execute("""
                    UPDATE Cargos
                    SET recrutamento = COALESCE(recrutamento, ?),
                        fonte_carga_horaria = COALESCE(fonte_carga_horaria, ?)
                    WHERE id = ?
                """, (recrut_clean, fonte_clean, cargo_id))
                atualizados += 1
                
        conn.commit()
        print(f"Sucesso: {atualizados} cargos atualizados com dados históricos do Excel!")
    else:
        print("Aviso: Planilha do Excel não encontrada localmente, dados históricos de Recrutamento e Fonte da CH não foram importados do Excel.")
        
    conn.close()
    print("Migração concluída com sucesso!")

if __name__ == "__main__":
    migrar()
